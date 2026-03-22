#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
╔══════════════════════════════════════════════════════════════════╗
║  DASHBOARD WEB — MUNICIPALES 2026 — 2ÈME TOUR                  ║
║  Communes de Haute-Garonne (31)                                 ║
║  Cabinet de la Présidente de Région                             ║
╚══════════════════════════════════════════════════════════════════╝

Lancer en local :
    streamlit run app.py

Sources de données (par priorité) :
    1. Scraping temps réel du site du Ministère de l'Intérieur
       resultats-elections.interieur.gouv.fr
    2. Fallback : fichiers Parquet data.gouv.fr (pipeline open data)
"""

import hashlib
import io
import re
from datetime import datetime
from zoneinfo import ZoneInfo

TZ_PARIS = ZoneInfo("Europe/Paris")

import pandas as pd
import requests
import streamlit as st
from bs4 import BeautifulSoup

# ─────────────────────────────────────────────────────────────────
# CONFIGURATION
# ─────────────────────────────────────────────────────────────────

ID_ELECTION = "2026_muni_t2"
CODE_DEPARTEMENT = "31"
CODE_REGION = "76"   # Occitanie

# Source 1 : site du Ministère de l'Intérieur (temps réel)
BASE_URL_INTERIEUR = (
    "https://www.resultats-elections.interieur.gouv.fr"
    "/municipales2026T2/ensemble_geographique"
    f"/{CODE_REGION}/{CODE_DEPARTEMENT}"
)

BASE_URL_INTERIEUR_T1 = (
    "https://www.resultats-elections.interieur.gouv.fr"
    "/municipales2026/ensemble_geographique"
    f"/{CODE_REGION}/{CODE_DEPARTEMENT}"
)

# Source 2 (fallback) : pipeline open data data.gouv.fr
# Ces fichiers Parquet sont mis à jour en continu pour chaque tour ;
# le filtre id_election == ID_ELECTION sélectionne automatiquement le bon tour.
URL_GENERAL   = "https://www.data.gouv.fr/api/1/datasets/r/ff16d511-10c0-405e-9b35-511723948fce"
URL_CANDIDATS = "https://www.data.gouv.fr/api/1/datasets/r/4d3b35f6-0b22-4415-a24c-419a676312e2"

COMMUNES_CIBLES = {
    "31555": {"nom": "Toulouse",                  "bv_ref": 263},
    "31149": {"nom": "Colomiers",                 "bv_ref": 25},
    "31557": {"nom": "Tournefeuille",             "bv_ref": 21},
    "31069": {"nom": "Blagnac",                   "bv_ref": 22},
    "31157": {"nom": "Cugnaux",                   "bv_ref": 15},
    "31506": {"nom": "Saint-Orens-de-Gameville",  "bv_ref": 12},
    "31561": {"nom": "L'Union",                   "bv_ref": 11},
    "31417": {"nom": "Pibrac",                    "bv_ref": 8},
    "31490": {"nom": "Saint-Jory",                "bv_ref": 4},
    "31182": {"nom": "Fenouillet",                "bv_ref": 4},
    "31395": {"nom": "Muret",                     "bv_ref": 19},
    "31424": {"nom": "Plaisance-du-Touch",        "bv_ref": 15},
    "31113": {"nom": "Castanet-Tolosan",          "bv_ref": 10},
    "31446": {"nom": "Ramonville-Saint-Agne",     "bv_ref": 11},
    "31483": {"nom": "Saint-Gaudens",             "bv_ref": 10},
    "31033": {"nom": "Auterive",                  "bv_ref": 6},
    "31291": {"nom": "Léguevin",                  "bv_ref": 8},
    "31169": {"nom": "Escalquens",                "bv_ref": 7},
    "31107": {"nom": "Carbonne",                  "bv_ref": 4},
    "31396": {"nom": "Nailloux",                  "bv_ref": 2},
    "31203": {"nom": "Frouzins",                  "bv_ref": 6},
    "31042": {"nom": "Bagnères-de-Luchon",        "bv_ref": 3},
    "31167": {"nom": "Encausse-les-Thermes",      "bv_ref": 1},
}

# ─────────────────────────────────────────────────────────────────
# SOURCE 1 : SCRAPING DU SITE DU MINISTÈRE (temps réel)
# ─────────────────────────────────────────────────────────────────

def _nettoyer_nombre(texte: str) -> int | None:
    """Convertit '1 211', '1\xa0211' ou '1\u202f211' en 1211."""
    s = texte.replace("\xa0", "").replace("\u202f", "").replace(" ", "").replace(",", "").strip()
    try:
        return int(s)
    except (ValueError, TypeError):
        return None


def scraper_commune(code_commune: str) -> dict | None:
    """Scrape les résultats d'une commune depuis le site du Ministère.
    Essaie l'URL T2 d'abord, puis fallback sur l'URL T1 (qui peut
    contenir les résultats T2 directement).
    Retourne None si résultats non parvenus ou erreur."""
    # Essayer T2 puis T1
    soup = None
    for base_url in [BASE_URL_INTERIEUR, BASE_URL_INTERIEUR_T1]:
        url = f"{base_url}/{code_commune}/"
        try:
            resp = requests.get(url, timeout=30, headers={
                "User-Agent": "DashboardMunicipales2026-HG31/1.0",
            })
            if resp.status_code == 404:
                continue
            resp.raise_for_status()
            resp.encoding = "utf-8"
            soup = BeautifulSoup(resp.text, "html.parser")
            break
        except Exception:
            continue

    if soup is None:
        return None

    # Détecter "résultats non parvenus"
    titre = soup.find("h5")
    if titre and "non parvenus" in titre.get_text().lower():
        return None

    # Détecter si la page contient des résultats de 2d tour
    page_text = soup.get_text()
    a_resultats_t2 = bool(re.search(r'2[de]\s+tour', page_text, re.IGNORECASE))

    # Détecter si dépouillement encore en cours ("incomplets calculés sur la base de X%")
    m = re.search(
        r'incomplets?\s+calcul[ée]s?\s+sur\s+la\s+base\s+de\s+([\d,\.]+)\s*%',
        page_text, re.IGNORECASE,
    )
    pct_depouille = float(m.group(1).replace(",", ".")) if m else None

    tables = soup.find_all("table")
    resultat = {"candidats": [], "participation": {}, "sieges_pourvus": 0, "sieges_a_pourvoir": 0, "pct_depouille": pct_depouille}

    # Si la page a des résultats T2 (page T1 mise à jour), il y a 2 blocs
    # de candidats et 2 de participation. On ne veut que le 1er de chaque
    # (= le T2). On utilise des compteurs pour s'arrêter.
    nb_tables_candidats = 0
    nb_tables_participation = 0
    max_tables = 1 if a_resultats_t2 else 99

    for table in tables:
        rows = table.find_all("tr")
        if not rows:
            continue

        # Identifier le type de table par l'en-tête
        header_cells = [c.get_text(strip=True).lower() for c in rows[0].find_all(["td", "th"])]

        # Table des sièges (header : '', 'Sièges à pourvoir', 'Sièges pourvus')
        if any("pourvoir" in h for h in header_cells) and not any("voix" in h for h in header_cells):
            for row in rows[1:]:
                cells = [c.get_text(strip=True) for c in row.find_all(["td", "th"])]
                if len(cells) >= 3:
                    a_pourvoir = _nettoyer_nombre(cells[1])
                    pourvus = _nettoyer_nombre(cells[2])
                    if a_pourvoir is not None:
                        resultat["sieges_a_pourvoir"] += a_pourvoir
                    if pourvus is not None:
                        resultat["sieges_pourvus"] += pourvus

        # Table de participation (contient "nombre", "% inscrits", etc.)
        elif any("nombre" in h for h in header_cells):
            if nb_tables_participation >= max_tables:
                continue
            nb_tables_participation += 1
            for row in rows[1:]:
                cells = [c.get_text(strip=True) for c in row.find_all(["td", "th"])]
                if len(cells) >= 2:
                    label = cells[0].lower()
                    if label in ("inscrits", "abstentions", "votants", "blancs", "nuls", "exprimés"):
                        val = _nettoyer_nombre(cells[1])
                        if val is not None:
                            resultat["participation"][label] = val

        # Table des candidatures (contient "voix", "conduite par", etc.)
        elif any("voix" in h for h in header_cells):
            if nb_tables_candidats >= max_tables:
                continue
            nb_tables_candidats += 1
            # Détecter si la colonne "Nuance" est présente (décale les indices)
            has_nuance = any("nuance" in h for h in header_cells)
            idx_voix = 3 if has_nuance else 2
            idx_pct_exp = 5 if has_nuance else 4
            idx_nuance = 2 if has_nuance else None
            for row in rows[1:]:
                cells = [c.get_text(strip=True) for c in row.find_all(["td", "th"])]
                if len(cells) > idx_pct_exp:
                    voix = _nettoyer_nombre(cells[idx_voix])
                    if voix is not None:
                        resultat["candidats"].append({
                            "liste": cells[0],
                            "candidat": cells[1],
                            "nuance": cells[idx_nuance] if idx_nuance else "",
                            "voix": voix,
                            "pct_exprimes": cells[idx_pct_exp].replace(",", "."),
                        })

    return resultat if resultat["participation"] else None


@st.cache_data(ttl=3600, show_spinner=False)
def charger_communes_t1_acquis() -> set:
    """Scrape les pages T1 (cache 1h) pour détecter les communes
    élues dès le 1er tour et qui n'ont donc pas de 2ème tour."""
    acquis = set()
    for code in COMMUNES_CIBLES:
        url = f"{BASE_URL_INTERIEUR_T1}/{code}/"
        try:
            resp = requests.get(url, timeout=15, headers={
                "User-Agent": "DashboardMunicipales2026-HG31/1.0",
            })
            resp.raise_for_status()
            resp.encoding = "utf-8"
            soup = BeautifulSoup(resp.text, "html.parser")
            page_text = soup.get_text()
            # Si la page mentionne des résultats de 2d tour, ce n'est PAS un T1 acquis
            if re.search(r'2[de]\s+tour', page_text, re.IGNORECASE):
                continue
            # Signal prioritaire : texte "pourvu au tour 1" dans le H5
            h5 = soup.find("h5")
            if h5 and re.search(r'pourvu\s+au\s+tour\s*1', h5.get_text(), re.IGNORECASE):
                acquis.add(code)
                continue
            sp, sa = 0, 0
            for table in soup.find_all("table"):
                rows = table.find_all("tr")
                if not rows:
                    continue
                hdrs = [c.get_text(strip=True).lower() for c in rows[0].find_all(["td", "th"])]
                if any("pourvoir" in h for h in hdrs) and not any("voix" in h for h in hdrs):
                    # Détecter les indices de colonnes depuis l'en-tête pour s'adapter
                    # à plusieurs structures possibles (2 ou 3 colonnes)
                    idx_ap = next((i for i, h in enumerate(hdrs) if "pourvoir" in h), None)
                    idx_pu = next(
                        (i for i, h in enumerate(hdrs) if "pourvus" in h and "pourvoir" not in h),
                        None,
                    )
                    for row in rows[1:]:
                        cells = [c.get_text(strip=True) for c in row.find_all(["td", "th"])]
                        if idx_ap is not None and idx_pu is not None and len(cells) > max(idx_ap, idx_pu):
                            # Structure avec labels en en-tête : (valeur_ap, valeur_pu)
                            a = _nettoyer_nombre(cells[idx_ap])
                            p = _nettoyer_nombre(cells[idx_pu])
                        elif len(cells) >= 3:
                            # Structure avec colonne label en col 0 : (label, ap, pu)
                            a = _nettoyer_nombre(cells[1])
                            p = _nettoyer_nombre(cells[2])
                        else:
                            continue
                        if a:
                            sa += a
                        if p:
                            sp += p
            # Fallback textuel si l'analyse du tableau n'a rien trouvé
            if sa == 0:
                page_text = soup.get_text()
                m_ap = re.search(r'(\d+)\s*sièges?\s*à\s*pourvoir', page_text, re.IGNORECASE)
                m_pu = re.search(r'(\d+)\s*sièges?\s*pourvus?', page_text, re.IGNORECASE)
                if m_ap and m_pu:
                    sa = int(m_ap.group(1))
                    sp = int(m_pu.group(1))
            if sa > 0 and sp >= sa:
                acquis.add(code)
        except Exception:
            pass
    return acquis


def _charger_depuis_interieur_impl() -> pd.DataFrame | None:
    """Scrape toutes les communes cibles depuis le site du Ministère.
    Retourne un DataFrame prêt à l'emploi, ou None si aucun résultat."""
    communes_t1_acquis = charger_communes_t1_acquis()
    lignes = []
    for code, info in sorted(COMMUNES_CIBLES.items(), key=lambda x: x[1]["nom"]):
        nom = info["nom"]

        # Toujours tenter le T2 d'abord
        try:
            res = scraper_commune(code)
        except Exception:
            res = None

        if res is None:
            # Pas de résultats T2 : vérifier si c'est parce que c'était acquis au T1
            statut_nd = "Élu(e) au 1er tour" if code in communes_t1_acquis else "résultats non parvenus"
            lignes.append({
                "Commune": nom, "Code_INSEE": code,
                "Votants": None, "Taux abstention (%)": None,
                "Candidat": "", "Nuance": "", "Liste": "",
                "Voix": None, "% exprimés": None,
                "Statut": statut_nd,
            })
            continue

        p = res["participation"]
        inscrits = p.get("inscrits", 0)
        votants = p.get("votants", 0)
        exprimes = p.get("exprimés", 0)
        abstentions = p.get("abstentions", 0)
        taux_abst = round(abstentions / inscrits * 100, 2) if inscrits > 0 else 0.0

        # Statut basé sur les sièges pourvus + avancement dépouillement
        sp = res["sieges_pourvus"]
        sa = res["sieges_a_pourvoir"]
        pct = res["pct_depouille"]
        if sa > 0 and sp >= sa:
            statut = "Élu(e)"
        elif pct is not None:
            statut = f"en cours ({pct:.1f}% dépouillé)"
        else:
            statut = "en attente"

        if not res["candidats"]:
            lignes.append({
                "Commune": nom, "Code_INSEE": code,
                "Votants": votants, "Taux abstention (%)": taux_abst,
                "Candidat": "(aucun candidat trouvé)", "Nuance": "", "Liste": "",
                "Voix": None, "% exprimés": None,
                "Statut": statut,
            })
            continue

        for c in sorted(res["candidats"], key=lambda x: x["voix"], reverse=True):
            try:
                pct = float(c["pct_exprimes"])
            except (ValueError, TypeError):
                pct = round(c["voix"] / exprimes * 100, 2) if exprimes > 0 else 0.0
            lignes.append({
                "Commune": nom,
                "Code_INSEE": code,
                "Votants": votants,
                "Taux abstention (%)": taux_abst,
                "Candidat": c["candidat"],
                "Nuance": c["nuance"],
                "Liste": c["liste"],
                "Voix": c["voix"],
                "% exprimés": pct,
                "Statut": statut,
            })

    if not lignes:
        return None
    return pd.DataFrame(lignes)


@st.cache_data(ttl=120, show_spinner=False)
def charger_depuis_interieur() -> pd.DataFrame | None:
    """Version cachée (2 min) de _charger_depuis_interieur_impl."""
    return _charger_depuis_interieur_impl()


# ─────────────────────────────────────────────────────────────────
# SOURCE 2 (FALLBACK) : FICHIERS PARQUET DATA.GOUV.FR
# ─────────────────────────────────────────────────────────────────

def recuperer_signatures_http() -> dict:
    """Requêtes HEAD légères pour détecter si les fichiers sources
    ont changé, sans télécharger le contenu."""
    signatures = {}
    for nom, url in [("general", URL_GENERAL), ("candidats", URL_CANDIDATS)]:
        try:
            resp = requests.head(url, timeout=15, allow_redirects=True)
            resp.raise_for_status()
            sig = (resp.headers.get("Last-Modified", "")
                   + "|" + resp.headers.get("Content-Length", "")
                   + "|" + resp.headers.get("ETag", ""))
            signatures[nom] = sig
        except Exception:
            signatures[nom] = None
    return signatures


def calculer_hash_donnees(df: pd.DataFrame) -> str:
    """Hash du contenu d'un DataFrame pour détecter des changements réels."""
    h = hashlib.md5(usedforsecurity=False)
    h.update(pd.util.hash_pandas_object(df).values.tobytes())
    return h.hexdigest()


@st.cache_data(ttl=300, show_spinner=False)
def charger_et_filtrer() -> tuple:
    """Télécharge les Parquet nationaux, filtre sur HG31 et retourne
    uniquement les DataFrames filtrés (légers).
    Cache 5 min — invalidé plus tôt si le HEAD détecte un changement.
    """
    resp_gen = requests.get(URL_GENERAL, timeout=120)
    resp_gen.raise_for_status()
    df_general = pd.read_parquet(io.BytesIO(resp_gen.content))
    del resp_gen

    resp_cand = requests.get(URL_CANDIDATS, timeout=120)
    resp_cand.raise_for_status()
    df_candidats = pd.read_parquet(io.BytesIO(resp_cand.content))
    del resp_cand

    gen = df_general[
        (df_general["id_election"] == ID_ELECTION) &
        (df_general["code_departement"] == CODE_DEPARTEMENT)
    ].copy()
    del df_general

    cand = df_candidats[
        (df_candidats["id_election"] == ID_ELECTION) &
        (df_candidats["code_departement"] == CODE_DEPARTEMENT)
    ].copy()
    del df_candidats

    return gen, cand


# ─────────────────────────────────────────────────────────────────
# TRAITEMENT DES DONNÉES (identique au script CLI)
# ─────────────────────────────────────────────────────────────────

def agreger_resultats(gen, cand):
    codes_cibles = set(COMMUNES_CIBLES.keys())

    gen_cibles  = gen[gen["code_commune"].isin(codes_cibles)].copy()
    cand_cibles = cand[cand["code_commune"].isin(codes_cibles)].copy()

    participation = gen_cibles.groupby("code_commune").agg(
        inscrits=("inscrits",    "sum"),
        abstentions=("abstentions", "sum"),
        votants=("votants",    "sum"),
        blancs=("blancs",     "sum"),
        nuls=("nuls",      "sum"),
        exprimes=("exprimes",   "sum"),
        nb_bv=("code_bv",   "nunique"),
    ).reset_index()

    participation["taux_abstention"] = (
        (participation["abstentions"]
         / participation["inscrits"].replace(0, float("nan"))) * 100
    ).round(2).fillna(0.0)

    candidat_cols = ["nom", "prenom"]
    if "liste"   in cand_cibles.columns: candidat_cols.append("liste")
    if "nuance"  in cand_cibles.columns: candidat_cols.append("nuance")

    group_cols = ["code_commune"] + candidat_cols
    resultats_cand = cand_cibles.groupby(group_cols, dropna=False).agg(
        voix=("voix", "sum"),
    ).reset_index()

    return participation, resultats_cand


def determiner_statut(code_commune, nb_bv_trouves):
    ref    = COMMUNES_CIBLES.get(code_commune, {})
    bv_ref = ref.get("bv_ref", 0)
    if bv_ref == 0:
        return "inconnu"
    elif nb_bv_trouves >= bv_ref:
        return "complet"
    else:
        return f"partiel ({nb_bv_trouves}/{bv_ref} BV)"


def construire_tableau_final(participation, resultats_cand, communes_t1_acquis=None):
    if communes_t1_acquis is None:
        communes_t1_acquis = set()
    lignes = []
    for code, info in sorted(COMMUNES_CIBLES.items(), key=lambda x: x[1]["nom"]):
        nom_commune = info["nom"]
        part = participation[participation["code_commune"] == code]

        if part.empty:
            statut_nd = "Élu(e) au 1er tour" if code in communes_t1_acquis else "données non disponibles"
            lignes.append({
                "Commune": nom_commune, "Code_INSEE": code,
                "Votants": None, "Taux abstention (%)": None,
                "Candidat": "", "Nuance": "", "Liste": "",
                "Voix": None, "% exprimés": None,
                "Statut": statut_nd,
            })
            continue

        row_p       = part.iloc[0]
        nb_votants  = int(row_p["votants"])
        nb_exprimes = int(row_p["exprimes"])
        taux_abst   = row_p["taux_abstention"]
        nb_bv       = int(row_p["nb_bv"])
        statut      = determiner_statut(code, nb_bv)

        cand_c = resultats_cand[resultats_cand["code_commune"] == code].copy()

        if cand_c.empty:
            lignes.append({
                "Commune": nom_commune, "Code_INSEE": code,
                "Votants": nb_votants, "Taux abstention (%)": taux_abst,
                "Candidat": "(aucun candidat trouvé)", "Nuance": "", "Liste": "",
                "Voix": None, "% exprimés": None,
                "Statut": statut,
            })
            continue

        cand_c["pct"] = (
            (cand_c["voix"] / nb_exprimes * 100).round(2)
            if nb_exprimes > 0 else 0
        )
        cand_c = cand_c.sort_values("voix", ascending=False)

        for _, r in cand_c.iterrows():
            lignes.append({
                "Commune": nom_commune,
                "Code_INSEE": code,
                "Votants": nb_votants,
                "Taux abstention (%)": taux_abst,
                "Candidat": f"{r.get('prenom', '')} {r.get('nom', '')}".strip(),
                "Nuance": r.get("nuance", "") or "",
                "Liste": r.get("liste", "") or "",
                "Voix": int(r["voix"]),
                "% exprimés": r["pct"],
                "Statut": statut,
            })

    return pd.DataFrame(lignes)


# ─────────────────────────────────────────────────────────────────
# EXPORT EXCEL EN MÉMOIRE (pour le bouton de téléchargement)
# ─────────────────────────────────────────────────────────────────

def generer_excel_bytes(df_final: pd.DataFrame) -> bytes:
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df_final.to_excel(writer, index=False, sheet_name="Résultats T2")
        ws = writer.sheets["Résultats T2"]

        # Colonnes à fusionner (celles liées à la commune, pas au candidat)
        cols_commune = ["Commune", "Code_INSEE", "Votants", "Taux abstention (%)", "Statut"]
        col_indices = {col: list(df_final.columns).index(col) + 1 for col in cols_commune}

        # Fusionner les cellules pour les communes avec plusieurs candidats
        row_idx = 2  # ligne 1 = en-tête
        while row_idx <= len(df_final) + 1:
            commune = ws.cell(row_idx, 1).value
            end_row = row_idx
            while end_row + 1 <= len(df_final) + 1 and ws.cell(end_row + 1, 1).value == commune:
                end_row += 1
            if end_row > row_idx:
                for col_name, col_idx in col_indices.items():
                    ws.merge_cells(
                        start_row=row_idx, start_column=col_idx,
                        end_row=end_row, end_column=col_idx,
                    )
                    ws.cell(row_idx, col_idx).alignment = Alignment(
                        vertical="center", wrap_text=True,
                    )
            row_idx = end_row + 1

        # Largeurs de colonnes
        for i, col in enumerate(df_final.columns, 1):
            col_lens = df_final[col].astype(str).str.len()
            max_len = max(len(str(col)), col_lens.max() if len(col_lens) > 0 else 0)
            ws.column_dimensions[ws.cell(1, i).column_letter].width = min(max_len + 3, 60)

        # Style en-tête
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # Bordures légères
        thin_border = Border(
            left=Side(style="thin", color="D9D9D9"),
            right=Side(style="thin", color="D9D9D9"),
            top=Side(style="thin", color="D9D9D9"),
            bottom=Side(style="thin", color="D9D9D9"),
        )
        for row in ws.iter_rows(min_row=2, max_row=len(df_final) + 1,
                                max_col=len(df_final.columns)):
            for cell in row:
                cell.border = thin_border

        # Feuille synthèse participation
        resume = df_final.drop_duplicates(subset=["Commune"])[
            ["Commune", "Code_INSEE", "Votants", "Taux abstention (%)", "Statut"]
        ].sort_values("Commune")
        resume.to_excel(writer, index=False, sheet_name="Synthèse participation")
        ws2 = writer.sheets["Synthèse participation"]
        for cell in ws2[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        for i, col in enumerate(resume.columns, 1):
            col_lens = resume[col].astype(str).str.len()
            max_len = max(len(str(col)), col_lens.max() if len(col_lens) > 0 else 0)
            ws2.column_dimensions[ws2.cell(1, i).column_letter].width = min(max_len + 3, 60)

    return buf.getvalue()


# ─────────────────────────────────────────────────────────────────
# INTERFACE STREAMLIT
# ─────────────────────────────────────────────────────────────────

def main():
    st.set_page_config(
        page_title="Municipales 2026 — Haute-Garonne",
        page_icon="🗳️",
        layout="wide",
    )

    st.title("🗳️ Municipales 2026 — 2ème tour — Haute-Garonne (31)")
    st.caption(
        "Source primaire : resultats-elections.interieur.gouv.fr · "
        "Fallback : data.gouv.fr · Rafraîchissement auto toutes les 2 min"
    )

    # ── Initialisation session state ───────────────────────────────
    if "hash_donnees" not in st.session_state:
        st.session_state.hash_donnees = None
    if "heure_maj" not in st.session_state:
        st.session_state.heure_maj = None
    if "donnees_nouvelles" not in st.session_state:
        st.session_state.donnees_nouvelles = False
    if "source_active" not in st.session_state:
        st.session_state.source_active = None

    # ── Rafraîchissement automatique toutes les 2 min ────────────────
    if "dernier_refresh" not in st.session_state:
        st.session_state.dernier_refresh = datetime.now(tz=TZ_PARIS)

    @st.fragment(run_every="120s")
    def auto_refresh():
        """Toutes les 120s : vide le cache et relance l'app complète."""
        maintenant = datetime.now(tz=TZ_PARIS)
        # Ne rerun que si au moins 60s se sont écoulées depuis le dernier refresh
        # (évite le rerun au tout premier rendu du fragment)
        ecart = (maintenant - st.session_state.dernier_refresh).total_seconds()
        if ecart >= 60:
            st.session_state.dernier_refresh = maintenant
            st.cache_data.clear()
            st.rerun(scope="app")

    auto_refresh()

    # ── Bandeau de statut ──────────────────────────────────────────
    heure = datetime.now(tz=TZ_PARIS).strftime("%H:%M:%S")
    heure_maj_str = (
        st.session_state.heure_maj.strftime("%d/%m/%Y à %H:%M:%S")
        if st.session_state.heure_maj
        else "—"
    )
    source = st.session_state.source_active or "—"
    if st.session_state.donnees_nouvelles:
        st.success(
            f"🆕 **Mise à jour détectée le {heure_maj_str} !** "
            f"Source : {source}"
        )
    else:
        st.info(
            f"🔄 Dernière vérification : {heure} · "
            f"Dernière MAJ : {heure_maj_str} · Source : {source}"
        )

    # ── Bouton de rafraîchissement manuel ──────────────────────────
    if st.button("🔄 Rafraîchir manuellement"):
        st.cache_data.clear()
        st.rerun()

    # ── Chargement des données (Ministère puis fallback data.gouv) ─
    df_final = None
    source = None

    # Source 1 : scraping Ministère de l'Intérieur
    try:
        with st.spinner("Récupération des résultats depuis le Ministère de l'Intérieur…"):
            df_final = charger_depuis_interieur()
        if df_final is not None and not df_final.empty:
            source = "Ministère de l'Intérieur (temps réel)"
    except Exception:
        df_final = None

    # Source 2 : fallback data.gouv.fr
    if df_final is None or df_final.empty:
        try:
            with st.spinner("Fallback : téléchargement depuis data.gouv.fr…"):
                gen, cand = charger_et_filtrer()
            if not gen.empty:
                participation, resultats_cand = agreger_resultats(gen, cand)
                df_final = construire_tableau_final(participation, resultats_cand, charger_communes_t1_acquis())
                source = "data.gouv.fr (Parquet)"
        except Exception as e:
            st.error(f"❌ Impossible de charger les données : {e}")
            st.stop()

    if df_final is None or df_final.empty:
        st.warning(
            "⚠️ Aucune donnée disponible. Les résultats ne sont pas encore publiés.\n\n"
            "Le dashboard vérifie automatiquement toutes les 2 minutes."
        )
        st.stop()

    st.session_state.source_active = source

    # ── Détection de changement réel (hash du contenu) ─────────────
    hash_actuel = calculer_hash_donnees(df_final)
    if hash_actuel != st.session_state.hash_donnees:
        if st.session_state.hash_donnees is not None:
            st.session_state.donnees_nouvelles = True
        st.session_state.hash_donnees = hash_actuel
        st.session_state.heure_maj = datetime.now(tz=TZ_PARIS)

    # ── Métriques globales ─────────────────────────────────────────
    nb_complets  = df_final[df_final["Statut"] == "Élu(e)"]["Commune"].nunique()
    nb_t1_acquis = df_final[df_final["Statut"] == "Élu(e) au 1er tour"]["Commune"].nunique()
    nb_en_cours  = df_final[df_final["Statut"].str.startswith("en cours", na=False)]["Commune"].nunique()
    nb_en_attente = df_final[df_final["Statut"] == "en attente"]["Commune"].nunique()
    nb_attente   = df_final[
        df_final["Statut"].isin(["données non disponibles", "résultats non parvenus"])
    ]["Commune"].nunique()

    c1, c2, c3, c4, c5, c6 = st.columns(6)
    c1.metric("Communes suivies",  len(COMMUNES_CIBLES))
    c2.metric("✅ Élu(e) T2",        nb_complets)
    c3.metric("🔵 Élu(e) T1",        nb_t1_acquis)
    c4.metric("🟠 En attente rés.",  nb_en_attente)
    c5.metric("⏳ En cours",          nb_en_cours)
    c6.metric("⚠️ Non parvenus",     nb_attente)

    st.divider()

    # ── Bouton téléchargement Excel (proéminent si MAJ) ────────────
    horodatage  = datetime.now(tz=TZ_PARIS).strftime("%Y%m%d_%H%M")
    excel_bytes = generer_excel_bytes(df_final)

    if st.session_state.donnees_nouvelles:
        st.download_button(
            label="📥 Télécharger le nouvel Excel (données mises à jour !)",
            data=excel_bytes,
            file_name=f"municipales_2026_T2_HG31_{horodatage}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary",
            on_click=lambda: st.session_state.update(donnees_nouvelles=False),
        )
    else:
        st.download_button(
            label="📥 Télécharger Excel",
            data=excel_bytes,
            file_name=f"municipales_2026_T2_HG31_{horodatage}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    st.divider()

    # ── Tableau résultats par commune ──────────────────────────────
    st.subheader("Résultats détaillés par commune")

    communes_liste = ["Toutes"] + sorted(df_final["Commune"].unique().tolist())
    choix = st.selectbox("Filtrer par commune", communes_liste)

    df_affiche = df_final if choix == "Toutes" else df_final[df_final["Commune"] == choix]

    def colorier_statut(val):
        if val == "Élu(e)":
            return "background-color: #d4edda; color: #155724;"
        elif val == "Élu(e) au 1er tour":
            return "background-color: #cce5ff; color: #004085;"
        elif val == "en attente":
            return "background-color: #fde8d0; color: #7d3c00;"
        elif str(val).startswith("en cours"):
            return "background-color: #fff3cd; color: #856404;"
        elif val in ("données non disponibles", "résultats non parvenus"):
            return "background-color: #f8d7da; color: #721c24;"
        return ""

    st.dataframe(
        df_affiche.style.map(colorier_statut, subset=["Statut"]),
        width="stretch",
        hide_index=True,
        height=600,
    )

    # ── Participation en un coup d'œil ─────────────────────────────
    st.subheader("Participation")
    synthese = df_final.drop_duplicates(subset=["Commune"])[[
        "Commune", "Votants", "Taux abstention (%)", "Statut"
    ]].sort_values("Taux abstention (%)").reset_index(drop=True)

    st.dataframe(
        synthese.style.map(colorier_statut, subset=["Statut"]),
        width="stretch",
        hide_index=True,
    )


if __name__ == "__main__":
    main()
